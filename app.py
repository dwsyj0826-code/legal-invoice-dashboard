"""
⚖️ 법무팀 외부 로펌 비용 대시보드
대웅제약 법무1팀 · 구글 드라이브 인보이스 자동 집계

구조: Google Drive(인보이스 PDF/xlsx) → 자동 파싱 → Streamlit 대시보드
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import pdfplumber
import openpyxl
import io
import re
import json
from datetime import datetime
from pathlib import Path

# ============================================================
# 페이지 설정
# ============================================================
st.set_page_config(
    page_title="정기자문 비용 현황",
    page_icon="⚖️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ============================================================
# 상수
# ============================================================

# 루트 폴더 ID (Streamlit secrets에서 가져오거나 기본값 사용)
ROOT_FOLDER_ID = st.secrets.get("ROOT_FOLDER_ID", "1HleAj4z6DH9KxMjyuf56lRD4b-c7pOyh")

# 2025년 폴더 ID (월별 폴더 구조, SLP·DLS·D&A만 파싱)
FOLDER_ID_2025 = st.secrets.get("FOLDER_ID_2025", "1ppBvKMRvHMgrvzQT5xcDuMR___mthgqk")

# 2025년 대상 로펌 (SLP·DLS·D&A만, 다른 로펌은 이 폴더에 없음)
FIRMS_2025 = {"SLP", "DLS", "D&A"}

# ★ 스냅샷 파일 경로 (GitHub 저장소 최상단에 두는 정적 데이터)
SNAPSHOT_PATH = Path(__file__).parent / "data_snapshot.json"

# ============================================================
# ★ 2025년 수동 입력 데이터 (VAT 제외 금액)
# ============================================================
# 2025년은 폴더 구조가 불규칙하고 파일 파싱 실패가 잦아 자동 파싱을 포기하고
# 수동 관리. 아래 dict의 숫자만 수정하면 대시보드에 즉시 반영됨.
# 금액 0 = 해당 월에 청구 없음 (표시 X).
# ============================================================
MANUAL_2025 = {
    "DLS": {
        1: 11_545_000,  2: 27_000_000,  3: 11_876_000,  4: 8_149_000,
        5:  3_885_000,  6:          0,  7:  3_208_000,  8:         0,
        9:  6_414_000, 10:  4_045_000, 11:  6_103_000, 12: 7_475_000,
    },
    "D&A": {
        1: 14_734_200,  2: 13_612_000,  3:  9_840_000,  4: 6_938_000,
        5:  5_000_000,  6:  5_000_000,  7:  5_000_000,  8: 5_000_000,
        9:  5_000_000, 10:  5_000_000, 11: 13_857_000, 12: 17_253_000,
    },
    "SLP": {
        1: 11_900_000,  2: 24_352_500,  3: 14_960_000,  4: 10_965_000,
        5:  4_122_500,  6: 13_387_500,  7:  6_205_000,  8:  4_037_500,
        9: 10_497_500, 10:  9_265_000, 11: 24_352_500, 12:  7_607_500,
    },
}

# 2025년 인보이스 링크 (Drive 파일 ID 매핑)
_DRIVE_VIEW = "https://drive.google.com/file/d/{}/view"
MANUAL_2025_LINKS = {
    "SLP": {
        1: _DRIVE_VIEW.format("1XgvZJYICrbep_ai6_uzw2hz9br-zBFjy"),
        2: _DRIVE_VIEW.format("1BuHbosIQSPUjcP9y9yDrGt4BVGecs-1o"),
        3: _DRIVE_VIEW.format("1x3tO9XFKb8dWsu090FYICbMAy_KKhVrE"),
        4: _DRIVE_VIEW.format("11ZorYxJBDlgOLpud2rT5QiKJLMrqUoA5"),
        5: _DRIVE_VIEW.format("1JE2vy90gxkE7EonIinDoJ0vNvuEh7lUA"),
        6: _DRIVE_VIEW.format("1AVDEC2txMz-lLRhvj42MuiUIHlo8j1Va"),
        7: _DRIVE_VIEW.format("1btB44kWkSuhGuBEkzrAAJkMr1I4oK4Ly"),
        8: _DRIVE_VIEW.format("1-Ec6ys_35gI5xK6KsRJVPk2JDElTOLjv"),
        9: _DRIVE_VIEW.format("1aNZ7h-Lz7VvpG6vrupzuLdW7uX8n7J5M"),
       10: _DRIVE_VIEW.format("1-nGKMV5lgEIWCeyh_DyKqsQUBhpgPqjw"),
       11: _DRIVE_VIEW.format("1T50PwmDiEZsgtD0gvN9Pi3DuhapFlwbZ"),
       12: _DRIVE_VIEW.format("1Uu37_coEQm6SxbzjzUFlfE0jCxbZ4aVC"),
    },
    "DLS": {
        1: _DRIVE_VIEW.format("1hqvTcIwuPNwKzryQEE4OWltRRrOZz_YS"),
        2: _DRIVE_VIEW.format("1jRroTg2A5mm_k1CSoTbT9JS_UJZbGXEe"),
        3: _DRIVE_VIEW.format("1BMTSywETWrflKNlmX2HLuq04-59zJa8p"),
        4: _DRIVE_VIEW.format("1GiXAF9KiBJavSEwOcBi69jIlejsIbMds"),
        # 5: DLS 5월분 링크 보류
        # 6, 8: DLS 청구 없음
        7: _DRIVE_VIEW.format("1DHihoOUe-1tZvfCaSvvgz3ZQnNZfQlQn"),
        9: _DRIVE_VIEW.format("1i1zIOBzmEBjF_cKWNACILy6yv5L9wKDU"),
       10: _DRIVE_VIEW.format("1YexQ6TTV6ZtbbR9PIWr9qaPtmiVcYXe_"),
       11: _DRIVE_VIEW.format("1uMKWo2853Up51qmX6fxpesJy9ZfSmQtJ"),
       12: _DRIVE_VIEW.format("16vtpoQvMOUnvnNOutcSXKEAgCn1Ix0Xm"),
    },
    "D&A": {
        1: _DRIVE_VIEW.format("1Y7YkOUVLxPf0V4qsUqZaOYf9SrFYuj4B"),
        2: _DRIVE_VIEW.format("1BknK5jttIfnSXfYeqsreQEh5SDUFs2NC"),
        3: _DRIVE_VIEW.format("1NPKPGfkAynFV7FOY2zp9dJXKaC8aR6Ji"),
        4: _DRIVE_VIEW.format("1cFWvRJgHmYQh4MQJM4VXxw3OZu2I6q7G"),
        5: _DRIVE_VIEW.format("1dm2BNuqoTCbxupkOmyrNeCEi03Yvu3JH"),
        6: _DRIVE_VIEW.format("1mztuMH_WZc-1KcfeqGFrPHk-NQ3rZWe9"),
        7: _DRIVE_VIEW.format("1nb9hj1FrAs99rXfpn5rusujgKZQlqnVF"),
        8: _DRIVE_VIEW.format("1iOrsG5_dURBS42D6WjBM1fhQgzLM3JEm"),
        9: _DRIVE_VIEW.format("1to_OChLoCtmmzSwgxUkIjOjehq9GKWi_"),
       10: _DRIVE_VIEW.format("18ewlRGslAHLGj3pKDPlSH6YN6lgSy3M3"),
       11: _DRIVE_VIEW.format("19ybE_b0QAQx2Z5RQI7JrmsOo9wfRyu8j"),
       12: _DRIVE_VIEW.format("12WqqLZTIlAc_Xuez7KcQ6SPenyZb90Dp"),
    },
}

# 2025년 인보이스 파일명 매핑 (상세 내역 표시용)
MANUAL_2025_FILENAMES = {
    "SLP": {
        1: "[SLP]2501_법무검토인보이스_고객양식.pdf",
        2: "2502_법무검토인보이스_고객양식.pdf",
        3: "2503_법무검토인보이스_SLP.pdf",
        4: "2504_법무검토인보이스_고객양식.pdf",
        5: "2505_법무검토인보이스_고객양식.pdf",
        6: "250715_법률자문료 INVOICE_6월분_대웅제약.pdf",
        7: "250818_법률자문료 INVOICE_7월분_대웅제약.pdf",
        8: "250916_법률자문료 INVOICE_8월분_대웅제약.pdf",
        9: "2509_법무검토인보이스_고객양식.pdf",
       10: "251119_법률자문료 INVOICE_10월분_대웅제약.pdf",
       11: "251217_법률자문료 INVOICE_11월분_대웅제약.pdf",
       12: "260116_법률자문료 INVOICE_12월분_대웅&대웅제약.pdf",
    },
    "DLS": {
        1: "[DLS]대웅제약 업무수행내역(2501)_합본_v2_20250220.xlsx",
        2: "대웅제약 업무수행내역(2502)(청구본)_DLS.xlsx",
        3: "대웅제약 업무수행내역(2503)_청구본_v2_DLS_20250423.xlsx",
        4: "대웅제약 업무수행내역(2504)_청구본_DLS _20250520.xlsx",
        7: "대웅제약 업무수행내역(2507)_DLS_청구_20250820.xlsx",
        9: "대웅제약 업무수행내역(2509)_DLS_청구본_20251023.xlsx",
       10: "대웅제약 업무수행내역(2510)_DLS_청구본.xlsx",
       11: "대웅제약 업무수행내역(2511)_DLS_청구본_20251229.xlsx",
       12: "대웅제약 업무수행내역(2512)청구본_DLS_SHL_20260202.xlsx",
    },
    "D&A": {
        1: "[대륙아주]250217_23-73009_자문_대웅제약.pdf",
        2: "250318_23-73009_자문_대웅제약.pdf",
        3: "250416_자문료 청구서_대륙아주.pdf",
        4: "250516_23-73009_자문_대웅제약_대륙아주.pdf",
        5: "250619 5월 자문료 청구서_대륙아주.pdf",
        6: "250715 6월 자문료 청구서(대웅제약).pdf",
        7: "250818 7월 자문료 청구서(대웅제약).pdf",
        8: "250911 8월 자문료 청구서(대웅제약).pdf",
        9: "251016 9월 자문료 청구서(대웅제약).pdf",
       10: "251118 10월 자문료 청구서(대웅제약).pdf",
       11: "251211 11월 자문료 청구서(대웅제약).pdf",
       12: "260120 12월 자문료 청구서(대웅제약).pdf",
    },
}

# 로펌별 차트 색상 (파스텔 톤, 통일감 있게)
FIRM_COLORS = {
    "광장": "#7BA7D9",       # 파스텔 블루
    "김앤장": "#E8998D",     # 파스텔 코랄
    "지평": "#A3C9A8",       # 파스텔 그린
    "D&A": "#B8A0D9",        # 파스텔 퍼플
    "율촌": "#F0C987",       # 파스텔 오렌지
    "율촌(영문)": "#E8D687",  # 파스텔 옐로우
    "SLP": "#9EC5D9",        # 파스텔 스카이
    "DLS": "#D9A87B",        # 파스텔 캐러멜
    "세종": "#A0D4C8",       # 파스텔 민트
    "세종_인도네시아": "#8FC7B9",  # 파스텔 틸
}

# 인보이스 파일 판별 키워드
INVOICE_KEYWORDS = ["자문료", "인보이스", "invoice", "청구", "보수금"]
EXCLUDE_KEYWORDS = ["할인 전", "할인전", "이체확인증", "납부확인서", "실비"]

# 폴더명 → 표시명 매핑
FIRM_DISPLAY_NAMES = {
    "율촌_일반 자문업무 (정성무, 유일한, 안유선)": "율촌",
    "율촌_일반 자문업무 (영문계약, 인니)": "율촌(영문)",
    "디엘에스": "DLS",       # 2025 7~12월 폴더용
    "대륙아주": "D&A",       # 2025 7~12월 폴더용
}


def extract_file_date(filename):
    """파일명에서 6자리 날짜(YYMMDD) 추출 → 정수 반환. 없으면 0."""
    m = re.search(r"(?<!\d)(2[56]\d{4})(?!\d)", filename)
    if m:
        return int(m.group(1))
    return 0


def extract_month_from_pdf_text(text):
    """
    PDF 본문에서 자문 수행 월 추출.
    광장: "보수금(2026 년 4 월)"
    율촌: "2026-05-01부터 2026-05-31까지"
    D&A: "1월 법률자문료" 또는 "(2026. 1.)"
    지평/김앤장/SLP: 유사 패턴
    """
    if not text:
        return None
    t = re.sub(r"\s+", " ", text)

    # 광장: "보수금(2026년 4월)" 또는 "보수금(2026 년 4 월)"
    m = re.search(r"보수금\s*\(\s*(202[4-9])\s*년\s*(\d{1,2})\s*월\s*\)", t)
    if m:
        return (int(m.group(1)), int(m.group(2)))

    # 율촌: "2026-05-01부터 2026-05-31까지"
    m = re.search(r"(202[4-9])-(\d{2})-\d{2}\s*부터\s*(202[4-9])-(\d{2})-\d{2}", t)
    if m:
        return (int(m.group(1)), int(m.group(2)))

    # "N월 법률자문료" / "N월 자문료" (D&A 스타일)
    m = re.search(r"(\d{1,2})\s*월\s*(?:법률\s*)?자문료", t)
    if m:
        return (2026, int(m.group(1)))

    # "(2026. N.)" / "(2026.N.)" (D&A Description Of Services 스타일)
    m = re.search(r"\(\s*(202[4-9])\.\s*(\d{1,2})\.\s*\)", t)
    if m:
        return (int(m.group(1)), int(m.group(2)))

    # 김앤장: "2026 년 5 월 1 일부터 2026 년 5 월 31 일까지"
    m = re.search(r"(202[4-9])\s*년\s*(\d{1,2})\s*월\s*\d+\s*일\s*부터", t)
    if m:
        return (int(m.group(1)), int(m.group(2)))

    # 지평/일반: "2026년 N월"
    m = re.search(r"(202[4-9])\s*년\s*(\d{1,2})\s*월", t)
    if m:
        return (int(m.group(1)), int(m.group(2)))

    return None


def prev_month(period_str):
    """(사용 안 함) '2026-01' → '2025-12'"""
    y, m = period_str.split("-")
    y, m = int(y), int(m)
    if m == 1:
        return f"{y-1}-12"
    return f"{y}-{m-1:02d}"


# ============================================================
# Google Drive 인증
# ============================================================

@st.cache_resource
def get_drive_service():
    """Google Drive API 서비스 생성 (서비스 계정 인증)"""
    creds_dict = json.loads(st.secrets["GOOGLE_SERVICE_ACCOUNT"])
    creds = service_account.Credentials.from_service_account_info(
        creds_dict,
        scopes=["https://www.googleapis.com/auth/drive.readonly"],
    )
    return build("drive", "v3", credentials=creds)


# ============================================================
# Google Drive 파일 조작
# ============================================================

def list_folder(service, folder_id):
    """폴더 내 모든 파일/하위폴더 목록"""
    items = []
    page_token = None
    while True:
        resp = (
            service.files()
            .list(
                q=f"'{folder_id}' in parents and trashed = false",
                fields="nextPageToken, files(id, name, mimeType, webViewLink)",
                pageToken=page_token,
                orderBy="name",
            )
            .execute()
        )
        items.extend(resp.get("files", []))
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return items


def download_file(service, file_id):
    """파일 바이트 다운로드"""
    req = service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    return buf


# ============================================================
# 파일 선택 로직
# ============================================================

def _has_keyword(filename):
    """인보이스 키워드 포함 여부"""
    low = filename.lower()
    return any(kw.lower() in low for kw in INVOICE_KEYWORDS)


def _is_excluded(filename):
    """제외 대상 여부"""
    return any(kw in filename for kw in EXCLUDE_KEYWORDS)


def select_invoices(files):
    """
    파일 목록에서 인보이스 후보 반환.
    '할인 전' 등 제외 키워드만 필터링하고, 모든 유효한 PDF/xlsx를 반환.
    같은 월+로펌 중복은 이후 파일명 날짜 기준으로 제거됨.
    """
    pdfs = [f for f in files if f["name"].lower().endswith(".pdf") and not _is_excluded(f["name"])]
    xlsxs = [f for f in files if f["name"].lower().endswith(".xlsx") and not _is_excluded(f["name"])]

    # PDF가 있으면 PDF, 없으면 xlsx
    if pdfs:
        return pdfs
    return xlsxs


# ============================================================
# 파일명에서 청구 월 추출
# ============================================================

def extract_month(filename):
    """파일명/폴더명 → (year, month) 또는 None"""

    # "N월" / "N월분"
    m = re.search(r"(\d{1,2})\s*월", filename)
    if m:
        month = int(m.group(1))
        y = re.search(r"(202[4-9])", filename)
        return (int(y.group(1)) if y else 2026, month)

    # "2026.06" / "2026. 06"
    m = re.search(r"(202[4-9])[.\s]+(\d{2})", filename)
    if m:
        return (int(m.group(1)), int(m.group(2)))

    # "(2604)", "(2510)" — DLS 스타일 (YY = 24~29)
    m = re.search(r"\((2[4-9])(\d{2})\)", filename)
    if m:
        return (2000 + int(m.group(1)), int(m.group(2)))

    # "26.04", "25.10" — 율촌 스타일
    m = re.search(r"(?<!\d)(2[4-9])\.(\d{2})", filename)
    if m:
        return (2000 + int(m.group(1)), int(m.group(2)))

    # "26MMDDDD", "25MMDDDD" — 율촌 ID 스타일 (B-26051222)
    m = re.search(r"(2[4-9])(\d{2})\d{4}", filename)
    if m:
        return (2000 + int(m.group(1)), int(m.group(2)))

    # "20260520"
    m = re.search(r"(202[4-9])(\d{2})\d{2}", filename)
    if m:
        return (int(m.group(1)), int(m.group(2)))

    return None


# ============================================================
# PDF / XLSX 금액 파싱
# ============================================================

def _clean(s):
    """금액 문자열 → 정수"""
    return int(re.sub(r"[^\d]", "", s))


def parse_pdf_amount(text, firm):
    """
    로펌별 PDF 텍스트에서 VAT 제외 최종 금액 추출.
    각 로펌 인보이스 양식에 맞춘 패턴 사용.
    """
    # 공백 정규화
    t = re.sub(r"\s+", " ", text)

    # --- 로펌별 패턴 ---

    if firm == "광장":
        # "보수금합계 : 19,976,550 원"
        m = re.search(r"보수금\s*합\s*계\s*[:：]\s*(?:금\s*)?([\d,]+)\s*원", t)
        if m:
            return _clean(m.group(1))

    elif firm == "김앤장":
        # "자 문 료 ￦ 3,000,000"
        m = re.search(r"자\s*문\s*료\s*[￦₩W]?\s*([\d,]+)", t)
        if m:
            return _clean(m.group(1))

    elif firm == "지평":
        # "소 계 ￦ 16,029,280"
        m = re.search(r"소\s*계\s*[￦₩W]?\s*([\d,]+)", t)
        if m:
            return _clean(m.group(1))

    elif firm == "D&A":
        # D&A 표준 양식: "공급가 W 18,880,000" (VAT 제외 정상가)
        # 여러 통화 기호 대응: W, ₩, ￦, \, ¥ 등
        m = re.search(r"공급가\s*[Ww₩￦\\￥]?\s*([\d,]{7,})", t)
        if m:
            return _clean(m.group(1))
        # 할인 있는 경우: "자문료 합계 (20% 할인 후) 34,269,600"
        m = re.search(r"자문료\s*합계\s*\(\s*\d+\s*%\s*할인\s*후\s*\)\s*([\d,]+)", t)
        if m:
            return _clean(m.group(1))
        # "자문료 합계" 단독
        m = re.search(r"자문료\s*합계\s*[:：]?\s*([\d,]{7,})", t)
        if m:
            return _clean(m.group(1))
        # "총 법률자문료(3+4) 20,768,000" 은 VAT 포함이라 마지막 순위
        m = re.search(r"총\s*법률자문료\s*\([^)]+\)\s*([\d,]+)", t)
        if m:
            return int(round(_clean(m.group(1)) / 1.1))

    elif firm.startswith("율촌"):
        # "1. 보수금 : 27,243,750원" (첫 번째 보수금, 보수금합계가 아닌 것)
        m = re.search(r"(?<!합)보수금\s*[:：]\s*([\d,]+)\s*원", t)
        if m:
            return _clean(m.group(1))

    elif firm == "SLP":
        # "법률 자문료 15,512,500"
        m = re.search(r"법률\s*자문료\s*([\d,]+)", t)
        if m:
            return _clean(m.group(1))
        # 새로운 SLP 양식 (2601_법무검토인보이스): TOTAL 컬럼 합산 필요
        # PDF에서는 표 파싱이 어려우니 마지막 큰 숫자 찾기
        amounts = re.findall(r"₩\s*([\d,]+)", t)
        if amounts:
            # 마지막 것이 총합계일 가능성이 높음
            try:
                total = _clean(amounts[-1])
                if total > 100_000:
                    return total
            except Exception:
                pass

    elif firm == "세종_인도네시아" or firm == "세종":
        # 세종은 띄어쓰기가 있고 원화기호가 \(백슬래시)로 표시됨
        # 라인 아이템: "법률고문료 \ 5,000,000" (VAT 제외)
        # 상단 문구: "...법률고문료 ₩ 5,500,000 (10% 부가가치세 포함)..." (VAT 포함, 제외 필요)
        t_ns = t.replace(" ", "")
        # 백슬래시 \ 뒤에 오는 법률고문료 금액 우선 (라인 아이템)
        m = re.search(r"법률고문료\\([\d,]+)", t_ns)
        if m:
            return _clean(m.group(1))
        # 백슬래시가 없으면 부가가치세 포함액에서 VAT 제거
        m = re.search(r"법률고문료\s*[₩￦W]?\s*([\d,]+)\s*\(10%\s*부가가치세\s*포함\)", t_ns)
        if m:
            return int(round(_clean(m.group(1)) / 1.1))
        # 기존 원 단위 패턴 (예비)
        m = re.search(r"청구\s*금액\s*[:：]?\s*([\d,]+)\s*원", t)
        if m:
            return _clean(m.group(1))
        m = re.search(r"자문료\s*합계\s*[:：]?\s*([\d,]+)\s*원", t)
        if m:
            return _clean(m.group(1))

    # --- 범용 폴백 ---
    # "소계", "공급가액", "보수금" 뒤 금액
    m = re.search(r"(?:소\s*계|공급가액?|보수금|자문료)\s*[￦₩W:：]?\s*([\d,]+)", t)
    if m:
        val = _clean(m.group(1))
        if val > 100_000:  # 최소 10만원 이상이어야 의미 있는 금액
            return val

    return None


def parse_xlsx_amount(buf):
    """
    DLS 등 xlsx 파일에서 금액 추출.
    TOTAL 컬럼 합산 또는 '총합계' 셀 값 사용.
    """
    try:
        wb = openpyxl.load_workbook(buf, data_only=True)
        best_total = 0

        for ws in wb.worksheets:
            # 템플릿 시트 스킵
            if "양식" in (ws.title or ""):
                continue

            # 방법 1: "총합계" 옆의 금액 찾기
            for row in ws.iter_rows(values_only=False):
                for i, cell in enumerate(row):
                    if cell.value and "총합계" in str(cell.value):
                        # 같은 행의 다른 셀에서 숫자 찾기
                        for other in row:
                            v = other.value
                            if isinstance(v, (int, float)) and v > 10_000:
                                best_total = max(best_total, int(v))
                            elif isinstance(v, str):
                                cleaned = re.sub(r"[^\d]", "", v)
                                if cleaned and int(cleaned) > 10_000:
                                    best_total = max(best_total, int(cleaned))

            # 방법 2: TOTAL 컬럼 합산
            if best_total == 0:
                header_row = None
                total_col = None
                for row in ws.iter_rows(values_only=False):
                    for cell in row:
                        if cell.value == "TOTAL":
                            header_row = cell.row
                            total_col = cell.column
                            break
                    if header_row:
                        break

                if header_row and total_col:
                    col_total = 0
                    for row in ws.iter_rows(
                        min_row=header_row + 1,
                        min_col=total_col,
                        max_col=total_col,
                    ):
                        v = row[0].value
                        if isinstance(v, (int, float)) and v > 0:
                            col_total += int(v)
                        elif isinstance(v, str):
                            cleaned = re.sub(r"[^\d]", "", v)
                            if cleaned:
                                col_total += int(cleaned)
                    if col_total > 0:
                        best_total = max(best_total, col_total)

        return best_total if best_total > 0 else None
    except Exception:
        return None


# ============================================================
# 데이터 수집 파이프라인
# ============================================================

def get_display_name(folder_name):
    """폴더명 → 대시보드 표시명"""
    return FIRM_DISPLAY_NAMES.get(folder_name, folder_name)


@st.cache_data(ttl=86400, show_spinner=False)
def collect_invoices():
    """
    전체 로펌 폴더를 스캔하여 인보이스 데이터 수집.
    반환: (records: list[dict], errors: list[str])
    """
    service = get_drive_service()
    records = []
    errors = []

    # 루트 폴더의 하위 폴더(= 로펌별 폴더) 조회
    firm_folders = [
        f for f in list_folder(service, ROOT_FOLDER_ID)
        if f["mimeType"] == "application/vnd.google-apps.folder"
    ]

    for folder in firm_folders:
        firm_raw = folder["name"]
        firm = get_display_name(firm_raw)
        contents = list_folder(service, folder["id"])

        # 파일과 하위폴더 분리
        sub_folders = [c for c in contents if c["mimeType"] == "application/vnd.google-apps.folder"]
        direct_files = [c for c in contents if c["mimeType"] != "application/vnd.google-apps.folder"]

        # 탐색 대상 그룹: (소스명, 파일목록)
        groups = []
        if direct_files:
            groups.append(("direct", direct_files))
        for sf in sub_folders:
            sf_files = [
                f for f in list_folder(service, sf["id"])
                if f["mimeType"] != "application/vnd.google-apps.folder"
            ]
            if sf_files:
                groups.append((sf["name"], sf_files))

        for source, files in groups:
            invoices = select_invoices(files)
            for inv in invoices:
                fname = inv["name"]

                # 파일명 우선 월 추출 (백업용)
                month_info_from_name = extract_month(fname)
                if not month_info_from_name and source != "direct":
                    month_info_from_name = extract_month(source)

                # 금액 + PDF 텍스트 기반 월 추출
                amount = None
                pdf_month_info = None
                try:
                    if fname.lower().endswith(".pdf"):
                        buf = download_file(service, inv["id"])
                        with pdfplumber.open(buf) as pdf:
                            text = "\n".join(
                                (p.extract_text() or "") for p in pdf.pages
                            )
                        # PDF 본문에서 자문 월 추출 (최우선)
                        pdf_month_info = extract_month_from_pdf_text(text)
                        amount = parse_pdf_amount(text, firm)
                    elif fname.lower().endswith(".xlsx"):
                        buf = download_file(service, inv["id"])
                        # DLS는 청구본 기준으로 파싱 (원본이 아닌 실제 청구액)
                        if firm == "DLS":
                            amount = _parse_dls_xlsx_2025(buf)
                        else:
                            amount = parse_xlsx_amount(buf)
                except Exception as e:
                    errors.append(f"[{firm}] 파싱 오류 ({fname}): {e}")

                # 월 결정: PDF 텍스트 → 파일명 순
                month_info = pdf_month_info or month_info_from_name
                if not month_info:
                    errors.append(f"[{firm}] 월 추출 실패: {fname}")
                    continue

                year, month = month_info

                if amount and amount > 0:
                    records.append(
                        {
                            "로펌": firm,
                            "연도": year,
                            "월": month,
                            "기간": f"{year}-{month:02d}",
                            "금액": amount,
                            "파일명": fname,
                            "링크": inv.get("webViewLink", ""),
                            "파일날짜": extract_file_date(fname),
                            "우선순위": (1 if ("최종" in fname or "할인 후" in fname) else 0),
                        }
                    )
                elif amount is None:
                    errors.append(f"[{firm}] 금액 추출 실패: {fname}")

    # 중복 제거: 같은 로펌+같은 월 → (1) '최종' 우선 (2) 파일명 날짜 최신 우선
    if records:
        df_tmp = pd.DataFrame(records)
        df_tmp = df_tmp.sort_values(["우선순위", "파일날짜"], ascending=[False, False])
        df_tmp = df_tmp.drop_duplicates(subset=["로펌", "기간"], keep="first")
        df_tmp = df_tmp.drop(columns=["우선순위"])
        records = df_tmp.to_dict("records")

    # 표시기간 = 실제 자문 수행 월 (PDF에서 추출된 그대로)
    for r in records:
        r["표시기간"] = r["기간"]
        r["표시연도"] = r["연도"]
        r["표시월"] = r["월"]

    return records, errors


# ============================================================
# ★ 2025년 파싱 섹션 (독립) ─ SLP·DLS·D&A 3개 로펌만
# ============================================================
# 이 섹션은 2025년 폴더 구조 전용입니다.
# 2025 폴더는 월별 서브폴더(25.2, 25.3, ..., 25.6)와
# "2025 7월~ 12월" 폴더로 구성되며, 파일명·서브폴더명으로
# 로펌을 식별합니다.
# ============================================================


def _parse_slp_pdf_2025(buf):
    """
    SLP PDF 라인 아이템 표에서 TOTAL 컬럼 합산 → 통합 총액 (VAT 제외).
    양식: [SLP]*_고객양식.pdf. 표에 CLIENT/DESC/DATE/ATTORNEY/HOURS/RATE/TOTAL/REMARKS 열.
    폴백: 표 인식 실패 시 텍스트에서 '총합계' 근처 값 또는 ₩숫자 패턴 합산.
    """
    # 우선 시도: pdfplumber extract_tables
    try:
        total = 0
        text_all = ""
        buf.seek(0)
        with pdfplumber.open(buf) as pdf:
            for page in pdf.pages:
                text_all += (page.extract_text() or "") + "\n"
                tables = page.extract_tables() or []
                for tbl in tables:
                    for row in tbl:
                        if not row or len(row) < 7:
                            continue
                        total_str = (row[6] or "").strip()
                        if not total_str or "TOTAL" in total_str.upper():
                            continue
                        m = re.match(r"^₩?\s*([\d,]+)$", total_str)
                        if m:
                            try:
                                total += int(re.sub(r"[^\d]", "", m.group(1)))
                            except ValueError:
                                pass
        if total > 0:
            return total

        # 폴백 1: 텍스트에서 '총합계' 근처 큰 값 (하단 요약 표)
        t = re.sub(r"\s+", " ", text_all)
        # "총합계 11,900,000" 또는 "총합계 ... 11,900,000 13,090,000"
        totals = re.findall(r"총합계[^\d]{1,20}([\d,]{7,})", t)
        cand_totals = []
        for tot in totals:
            v = int(re.sub(r"[^\d]", "", tot))
            if v > 100_000:
                cand_totals.append(v)
        if cand_totals:
            # VAT 제외가 우선 (작은 값, 큰 값은 VAT 포함일 것)
            return min(cand_totals)

        # 폴백 2: 라인 아이템 ₩숫자 다 합산 (인식되는 표만)
        amounts = re.findall(r"₩\s*([\d,]+)", t)
        if amounts:
            total = 0
            for a in amounts[:-2]:  # 마지막 몇 개는 요약 총합일 가능성 - 라인만 합산
                v = int(re.sub(r"[^\d]", "", a))
                if v > 1000:  # 노이즈 필터
                    total += v
            if total > 100_000:
                return total

        return None
    except Exception:
        return None


def _parse_draju_pdf_2025(buf):
    """
    D&A(대륙아주) PDF: VAT 제외 통합 청구액.
    양식 1 (2025 후반~2026): "공급가 W 14,734,200"
    양식 2 (2025년 상반기): 표지가 이미지 → 페이지 2 "자문료 합계(3) 14,734,200"
    """
    try:
        with pdfplumber.open(buf) as pdf:
            text = "\n".join((p.extract_text() or "") for p in pdf.pages)
        t = re.sub(r"\s+", " ", text)

        # 양식 1: 공급가
        m = re.search(r"공급가\s*[Ww₩￦\\￥]?\s*([\d,]{7,})", t)
        if m:
            return int(re.sub(r"[^\d]", "", m.group(1)))

        # 양식 2: 자문료 합계 or 자문료 합계(N)
        m = re.search(r"자문료\s*합계\s*(?:\(\d+\))?\s*[:：]?\s*([\d,]{7,})", t)
        if m:
            return int(re.sub(r"[^\d]", "", m.group(1)))

        return None
    except Exception:
        return None


def _parse_dls_xlsx_2025(buf):
    """
    DLS xlsx: '청구본' (할인 적용 최종 청구액) 통합 총액.
    다양한 시트 구조 대응:
    - 합본 파일 (2501): 원본/청구본 시트 병존
    - 청구본 단일 파일 (2502~): 시트 하나 또는 데이터+rate 두 시트
    전략: 모든 (양식 아닌) 시트를 스캔, 마지막 '총합계' 행의 값들 중
          두 값이면 작은 값(청구본), 하나면 그 값, 다수면 최댓값과 최솟값 중
          최솟값 (할인 적용된 값).
    폴백: '최종 청구금액' 컬럼 헤더 아래 값, TOTAL 컬럼 합산.
    """
    try:
        wb = openpyxl.load_workbook(buf, data_only=True)

        # 우선순위 정렬: 청구본 rate > 청구본 > rate > 그 외
        def sheet_priority(ws):
            t = (ws.title or "")
            tl = t.lower()
            if "양식" in t:
                return -1  # 스킵
            score = 0
            if "청구본" in t: score += 4
            if "rate" in tl: score += 2
            return score

        sheets = [ws for ws in wb.worksheets if sheet_priority(ws) >= 0]
        sheets.sort(key=sheet_priority, reverse=True)

        # 각 시트에서 '총합계'/'총 금액' 셀 뒤의 값만 추출
        # (같은 행에 정경태 450000 같은 rate 값이 있어도 무시)
        def find_totals_in_sheet(ws):
            last_totals = []
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                # 이 행에서 "총합계" 또는 "총 금액" 셀 위치 찾기
                total_positions = [
                    i for i, v in enumerate(row)
                    if v and ("총합계" in str(v) or "총 금액" in str(v))
                ]
                if not total_positions:
                    continue
                # 마지막 "총합계" 셀 오른쪽 값들만 수집
                last_pos = total_positions[-1]
                right_vals = [
                    float(v) for v in row[last_pos + 1:]
                    if isinstance(v, (int, float)) and v > 100_000
                ]
                if right_vals:
                    last_totals = right_vals  # 마지막 발견된 총합계 행 값
            return last_totals

        for ws in sheets:
            last_total_nums = find_totals_in_sheet(ws)
            if last_total_nums:
                if len(last_total_nums) >= 2:
                    # 두 값 이상: 청구본(할인 적용, 작은 값) 선택
                    return int(round(min(last_total_nums)))
                return int(round(last_total_nums[0]))

        # 폴백 1: '최종 청구금액' 헤더 아래 열 합산
        for ws in sheets:
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    if cell.value and "최종 청구금액" in str(cell.value):
                        col = cell.column
                        col_total = 0
                        for r in ws.iter_rows(min_row=cell.row + 1,
                                              min_col=col, max_col=col,
                                              values_only=True):
                            v = r[0]
                            if isinstance(v, (int, float)) and v > 10_000:
                                col_total += v
                        if col_total > 100_000:
                            return int(round(col_total))

        # 폴백 2: TOTAL 컬럼 합산 (라인 아이템)
        for ws in sheets:
            header_row = header_col = None
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    if cell.value and str(cell.value).strip().upper() == "TOTAL":
                        header_row = cell.row
                        header_col = cell.column
                        break
                if header_row:
                    break
            if header_row and header_col:
                col_total = 0
                for r in ws.iter_rows(min_row=header_row + 1,
                                      min_col=header_col, max_col=header_col,
                                      values_only=True):
                    v = r[0]
                    if isinstance(v, (int, float)) and v > 0:
                        col_total += v
                    elif isinstance(v, str):
                        cleaned = re.sub(r"[^\d]", "", v)
                        if cleaned:
                            try:
                                col_total += int(cleaned)
                            except ValueError:
                                pass
                if col_total > 100_000:
                    return int(round(col_total))

        return None
    except Exception:
        return None


def _extract_month_from_2025_folder(folder_name):
    """
    2025 월별 폴더명에서 자문 수행 월 추출.
    예: '25. 2.정기자문료(25년1월분)' → (2025, 1)
        '25. 6. 정기자문료(25년 5월분)' → (2025, 5)
    """
    m = re.search(r"25\s*년\s*(\d{1,2})\s*월\s*분", folder_name)
    if m:
        return (2025, int(m.group(1)))
    # 24년 12월분 (25.1 폴더) 등 제외 대상
    m = re.search(r"24\s*년\s*(\d{1,2})\s*월\s*분", folder_name)
    if m:
        return (2024, int(m.group(1)))
    return None


@st.cache_data(ttl=86400, show_spinner=False)
def collect_invoices_2025():
    """
    2025년 데이터: 수동 입력 (자동 파싱 X).
    MANUAL_2025 dict의 값을 records 형식으로 변환하여 반환.
    금액이 0이거나 None이면 스킵 (해당 월 청구 없음).
    """
    records = []
    for firm, months in MANUAL_2025.items():
        for month, amount in months.items():
            if not amount or amount <= 0:
                continue
            link = MANUAL_2025_LINKS.get(firm, {}).get(month, "")
            fname = MANUAL_2025_FILENAMES.get(firm, {}).get(
                month, f"2025-{month:02d} {firm} 정기자문료"
            )
            records.append({
                "로펌": firm,
                "연도": 2025, "월": month,
                "기간": f"2025-{month:02d}",
                "표시기간": f"2025-{month:02d}",
                "표시연도": 2025, "표시월": month,
                "금액": amount,
                "파일명": fname,
                "링크": link,
            })
    return records, []


# ============================================================
# ★ 스냅샷 (JSON 정적 캐시) 로드/저장
# ============================================================
# 확정된 기간의 데이터를 GitHub 저장소의 data_snapshot.json으로 보관.
# 앱은 이 JSON을 즉시 로드하여 파싱 없이 표시.
# 스냅샷 이후 기간(=아직 확정 안 된 최신 분기)만 라이브로 파싱.
# ============================================================


def load_snapshot():
    """
    data_snapshot.json 로드. 파일이 없으면 None.
    반환: {"confirmed_until": "YYYY-MM", "records": [...], "generated_at": "..."}
    """
    if not SNAPSHOT_PATH.exists():
        return None
    try:
        with open(SNAPSHOT_PATH, encoding="utf-8") as f:
            data = json.load(f)
        # 기본 필드 검증
        if not isinstance(data, dict) or "records" not in data:
            return None
        return data
    except Exception:
        return None


def _admin_snapshot_ui():
    """
    관리자 스냅샷 생성 UI (사이드바 expander 안에 배치).
    확정 기준월 입력 → 전체 파싱 → 필터링 → JSON 다운로드.
    """
    st.caption(
        "확정할 기준월까지의 데이터를 JSON으로 만들어 다운로드합니다.\n"
        "다운로드된 파일을 GitHub 저장소 최상단에 업로드하시면 됩니다."
    )

    confirm_until = st.text_input(
        "확정 기준월 (YYYY-MM)",
        value="",
        placeholder="예: 2026-03",
        help="이 월(포함) 이하의 모든 데이터를 확정 처리합니다. "
             "예: 2026-03 입력 시 2025 전체 + 2026 1분기(1~3월)까지 확정.",
        key="snapshot_confirm_until",
    )

    if st.button("스냅샷 생성", key="snapshot_gen", type="primary"):
        if not re.match(r"^\d{4}-\d{2}$", (confirm_until or "").strip()):
            st.error("형식이 잘못됨. 예: 2026-03")
            return

        target = confirm_until.strip()
        with st.spinner("전체 인보이스 파싱 중... (수 분 소요)"):
            # 캐시 무시하고 최신 상태로 파싱
            st.cache_data.clear()
            records_25, errors_25 = collect_invoices_2025()
            records_26, errors_26 = collect_invoices()

            all_records = records_25 + records_26
            # 표시기간이 확정 기준월 이하인 것만 스냅샷에 포함
            filtered = [r for r in all_records
                        if str(r.get("표시기간", "")) <= target]

        snapshot = {
            "confirmed_until": target,
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "record_count": len(filtered),
            "records": filtered,
        }

        st.success(
            f"✅ 확정 대상 {len(filtered)}건 파싱 완료 (2025년 이후 ~ {target})"
        )

        # 로펌별 요약 (검증용)
        firms_summary = {}
        for r in filtered:
            firms_summary[r["로펌"]] = firms_summary.get(r["로펌"], 0) + 1
        st.caption(
            "· 로펌별: " + ", ".join(f"{k}({v}건)" for k, v in sorted(firms_summary.items()))
        )

        # 파싱 이슈가 있으면 표시
        all_errors = errors_25 + errors_26
        if all_errors:
            with st.expander(f"⚠️ 파싱 이슈 {len(all_errors)}건"):
                for e in all_errors:
                    st.caption(e)

        st.download_button(
            "📥 data_snapshot.json 다운로드",
            data=json.dumps(snapshot, ensure_ascii=False, indent=2, default=str),
            file_name="data_snapshot.json",
            mime="application/json",
            key="snapshot_dl",
        )
        st.caption(
            "📌 다운로드 후: GitHub 저장소에서 이 파일을 "
            "`data_snapshot.json` 으로 저장(덮어쓰기) → commit → Streamlit 자동 재배포"
        )


# ============================================================
# 대시보드 UI
# ============================================================

def apply_custom_css():
    """대시보드 커스텀 스타일"""
    st.markdown(
        """
        <style>
        /* 전체 폰트 크기 */
        html, body, [class*="css"] {
            font-size: 16px;
        }
        /* KPI 카드 */
        div[data-testid="stMetric"] {
            background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
            border-radius: 12px;
            padding: 16px 20px;
            border-left: 4px solid #1B4F72;
        }
        div[data-testid="stMetric"] label {
            font-size: 0.95rem !important;
            color: #495057;
        }
        div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
            font-size: 1.6rem !important;
            font-weight: 700;
            color: #1B4F72;
        }
        /* 사이드바 */
        section[data-testid="stSidebar"] {
            background: #f8f9fa;
        }
        /* ★★ 사이드바 상단 헤더/공백 완전 제거 ★★ */
        [data-testid="stSidebarHeader"] {
            display: none !important;
        }
        [data-testid="stSidebar"] > div:first-child {
            padding-top: 0.5rem !important;
        }
        [data-testid="stSidebarUserContent"] {
            padding-top: 0.5rem !important;
            margin-top: 0 !important;
        }
        [data-testid="stSidebar"] .block-container {
            padding-top: 0.5rem !important;
        }
        /* collapse 버튼 위치 */
        [data-testid="stSidebarCollapseButton"] {
            top: 0.5rem !important;
        }
        /* multiselect 태그 색상 (빨강 → 진한 회색) */
        span[data-baseweb="tag"] {
            background-color: #5a5a5a !important;
        }
        span[data-baseweb="tag"] span {
            color: white !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def check_password():
    """비밀번호 인증. 통과 시 True 반환."""
    if st.session_state.get("password_correct", False):
        return True

    expected = st.secrets.get("DASHBOARD_PASSWORD", "dw2026")

    st.title("🔒 정기자문 비용 현황")
    st.caption("대웅제약 법무1팀 · 접근 인증")
    st.markdown("")

    pw = st.text_input("비밀번호", type="password", key="pw_input")
    if st.button("확인", type="primary"):
        if pw == expected:
            st.session_state["password_correct"] = True
            st.rerun()
        else:
            st.error("비밀번호가 일치하지 않습니다.")
    return False


def main():
    if not check_password():
        st.stop()

    apply_custom_css()

    st.title("⚖️ 정기자문 비용 현황")
    st.caption("대웅제약 법무1팀 · 구글 드라이브 인보이스 자동 집계")

    # ---- 데이터 로드 (스냅샷 + 라이브) ----
    snapshot = load_snapshot()

    if snapshot:
        confirmed_until = snapshot.get("confirmed_until", "")
        confirmed_records = snapshot.get("records", [])
        with st.spinner(f"📂 최신 데이터 (`{confirmed_until}` 이후) 파싱 중..."):
            # 스냅샷 이후 기간만 라이브 파싱
            live_records, errors = collect_invoices()
            live_records = [
                r for r in live_records
                if str(r.get("표시기간", "")) > confirmed_until
            ]
        records = confirmed_records + live_records
        st.caption(
            f"✅ 확정 데이터: **{confirmed_until}** 까지 "
            f"({len(confirmed_records)}건, "
            f"{snapshot.get('generated_at', '')[:10]} 생성) · "
            f"라이브 데이터: {len(live_records)}건"
        )
    else:
        with st.spinner("📂 구글 드라이브에서 인보이스를 읽는 중..."):
            records_2026, errors_2026 = collect_invoices()
            records_2025, errors_2025 = collect_invoices_2025()
            records = records_2026 + records_2025
            errors = errors_2026 + errors_2025
        st.info("💡 스냅샷 파일이 없어 전체 라이브 파싱합니다. 사이드바 '🛠 관리자'에서 스냅샷을 생성하세요.")

    if not records:
        st.error("인보이스 데이터를 찾을 수 없습니다. 폴더 공유 설정을 확인해 주세요.")
        if errors:
            with st.expander("⚠️ 오류 상세"):
                for e in errors:
                    st.text(e)
        st.stop()

    df = pd.DataFrame(records)

    # ---- 사이드바 필터 ----
    with st.sidebar:
        # 대시보드 제목·부제
        st.markdown(
            """
            <div style="padding:8px 0 4px 0; border-bottom:1px solid #dee2e6; margin-bottom:12px;">
                <div style="font-size:20px; font-weight:700; color:#1B4F72;">⚖️ 정기자문 비용</div>
                <div style="font-size:13px; color:#6c757d;">대웅제약 법무1팀</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.header("📊 조회 설정")

        # ═════════ [1] 표시 방식 (차트·KPI 표현) ═════════
        st.markdown("#### 🎨 표시 방식")

        view_mode = st.radio(
            "조회 모드",
            ["단일 연도", "두 연도 비교"],
            help="단일 연도: 로펌별 막대그래프. 두 연도 비교: 연도별 라인 겹쳐 표시.",
            key="view_mode_widget",
        )

        period_unit = st.radio(
            "집계 단위",
            ["월별", "분기별", "반기별", "연도별"],
            help="차트 X축·KPI(평균/최고 지출) 기준. 데이터를 어떤 단위로 묶어 볼지 결정.",
        )

        # 차트 유형 (단일 연도 모드에서만 도넛 옵션 제공)
        if view_mode == "단일 연도":
            chart_type = st.radio(
                "차트 유형",
                ["시계열 (막대)", "비중 (도넛)"],
                help="시계열: 월/분기 흐름 파악. 비중: 로펌별 비율 파악.",
            )
        else:
            chart_type = "시계열 (막대)"

        if view_mode == "두 연도 비교" and period_unit == "연도별":
            st.caption("💡 연도별 + 두 연도 비교: 각 연도당 값이 1개뿐이라 라인 형태가 안 나옴.")

        st.divider()

        # ═════════ [2] 데이터 필터 (전체 대시보드에 적용) ═════════
        st.markdown("#### 🔍 데이터 필터")

        years = sorted(df["표시연도"].unique())

        if view_mode == "단일 연도":
            sel_year = st.selectbox("연도", years, index=len(years) - 1, key="sel_year_widget")
            compare_years = None
        else:
            if len(years) < 2:
                st.warning("2개 이상의 연도 데이터가 필요합니다.")
                compare_years = years
                sel_year = years[0]
            else:
                compare_years = st.multiselect(
                    "비교할 연도 (2개 이상)",
                    years,
                    default=years[-2:],
                    key="compare_years_widget",
                )
                if len(compare_years) < 2:
                    st.warning("최소 2개 연도를 선택해주세요.")
                sel_year = None

        all_firms = sorted(df["로펌"].unique())

        if "sel_firms_state" not in st.session_state:
            st.session_state["sel_firms_state"] = all_firms

        bc1, bc2 = st.columns(2)
        if bc1.button("전체 선택", use_container_width=True):
            st.session_state["sel_firms_state"] = all_firms
            st.rerun()
        if bc2.button("전체 해제", use_container_width=True):
            st.session_state["sel_firms_state"] = []
            st.rerun()

        sel_firms = st.multiselect(
            "로펌",
            all_firms,
            default=st.session_state["sel_firms_state"],
            key="sel_firms_widget",
        )
        st.session_state["sel_firms_state"] = sel_firms

        st.divider()

        # ═════════ [3] 기타 ═════════
        st.markdown("#### ⚙️ 기타")
        if st.button("🔄 새로고침", use_container_width=True, help="Google Drive에서 최신 인보이스 다시 파싱"):
            st.cache_data.clear()
            st.rerun()

        if errors:
            with st.expander(f"⚠️ 파싱 이슈 ({len(errors)}건)"):
                for e in errors:
                    st.caption(e)

        # 관리자 스냅샷 UI (분기마다 한 번 사용)
        with st.expander("🛠 관리자 스냅샷 (분기 확정용)"):
            _admin_snapshot_ui()

    # ---- 필터 적용 ----
    if view_mode == "단일 연도":
        fdf = df[(df["표시연도"] == sel_year) & (df["로펌"].isin(sel_firms))].copy()
    else:
        fdf = df[(df["표시연도"].isin(compare_years)) & (df["로펌"].isin(sel_firms))].copy()

    if fdf.empty:
        st.warning("선택한 조건에 해당하는 데이터가 없습니다.")
        st.stop()

    # ---- 기간 집계 키 생성 ----
    def build_agg_key(r, unit, cross_year=False):
        if unit == "월별":
            return f"{r['표시월']}월" if cross_year else r["표시기간"]
        elif unit == "분기별":
            q = (r["표시월"] - 1) // 3 + 1
            return f"Q{q}" if cross_year else f"{r['표시연도']}-Q{q}"
        elif unit == "반기별":
            h = "상반기" if r["표시월"] <= 6 else "하반기"
            return h if cross_year else f"{r['표시연도']}-{h}"
        else:
            return str(r["표시연도"])

    def sort_key(row):
        if period_unit == "월별":
            return row["표시월"]
        elif period_unit == "분기별":
            return (row["표시월"] - 1) // 3 + 1
        elif period_unit == "반기별":
            return 1 if row["표시월"] <= 6 else 2
        else:
            return int(row["표시연도"])

    cross_year = (view_mode == "두 연도 비교")
    agg = fdf.groupby(["로펌", "표시연도", "표시월", "표시기간"], as_index=False)["금액"].sum()
    agg["집계"] = agg.apply(lambda r: build_agg_key(r, period_unit, cross_year), axis=1)
    agg["__sort"] = agg.apply(sort_key, axis=1)
    chart_df = agg.groupby(["집계", "로펌"], as_index=False)["금액"].sum()

    # ---- KPI (집계 단위 반응형) ----
    unit_labels = {
        "월별": ("월평균", "최고 지출월"),
        "분기별": ("분기평균", "최고 지출분기"),
        "반기별": ("반기평균", "최고 지출반기"),
        "연도별": ("연도평균", "최고 지출연도"),
    }
    avg_label, peak_label = unit_labels[period_unit]

    if period_unit == "월별":
        period_key = fdf["표시기간"]
    elif period_unit == "분기별":
        period_key = fdf.apply(lambda r: f"{r['표시연도']}-Q{(r['표시월']-1)//3+1}", axis=1)
    elif period_unit == "반기별":
        period_key = fdf.apply(
            lambda r: f"{r['표시연도']}-{'상반기' if r['표시월'] <= 6 else '하반기'}", axis=1
        )
    else:
        period_key = fdf["표시연도"].astype(str)

    period_totals = fdf.assign(__pk=period_key).groupby("__pk")["금액"].sum()

    # 조회 기간 (실제 데이터가 있는 범위)
    all_periods_sorted = sorted(fdf["표시기간"].unique())
    period_from = all_periods_sorted[0] if all_periods_sorted else "-"
    period_to = all_periods_sorted[-1] if all_periods_sorted else "-"
    n_periods = len(period_totals)  # 집계 단위별 몇 개 구간?

    # ★ 조회 기간 배너
    st.markdown(
        f"""
        <div style="background:#eaf3fb; padding:10px 16px; border-radius:8px;
                    border-left:4px solid #1B4F72; margin-bottom:12px; font-size:15px;">
            📅 <b>조회 기간</b>: {period_from} ~ {period_to}
            &nbsp;·&nbsp; 집계 단위: <b>{period_unit}</b>
            &nbsp;·&nbsp; 총 {n_periods}개 {period_unit.replace('별','')}
        </div>
        """,
        unsafe_allow_html=True,
    )

    # 5번째 KPI: 직전 기간 대비 (단일) 또는 전년 대비 (비교)
    delta_label = None
    delta_value = None
    delta_pct = None
    delta_help = None

    if view_mode == "단일 연도":
        unit_prev = {
            "월별": "전월", "분기별": "전분기",
            "반기별": "전반기", "연도별": "전년"
        }[period_unit]
        delta_label = f"{unit_prev} 대비"
        sorted_series = period_totals.sort_index()
        if len(sorted_series) >= 2:
            last_v = float(sorted_series.iloc[-1])
            prev_v = float(sorted_series.iloc[-2])
            delta_value = f"₩{last_v:,.0f}"
            if prev_v > 0:
                pct = (last_v - prev_v) / prev_v * 100
                delta_pct = f"{pct:+.1f}%"
                delta_help = f"최신 {period_unit.replace('별','')} 값이 직전 {period_unit.replace('별','')} 대비 얼마나 변했는지"
            else:
                delta_pct = "N/A"
        else:
            delta_value = "-"
            delta_help = f"직전 {period_unit.replace('별','')} 데이터가 없어 계산 불가"
    else:  # 두 연도 비교
        delta_label = "전년 대비 총액"
        year_totals = fdf.groupby("표시연도")["금액"].sum().sort_index()
        if len(year_totals) >= 2:
            latest_y = float(year_totals.iloc[-1])
            prev_y = float(year_totals.iloc[-2])
            delta_value = f"₩{latest_y:,.0f}"
            if prev_y > 0:
                pct = (latest_y - prev_y) / prev_y * 100
                delta_pct = f"{pct:+.1f}%"
                delta_help = f"{int(year_totals.index[-1])}년 총액이 {int(year_totals.index[-2])}년 대비 얼마나 변했는지"
            else:
                delta_pct = "N/A"
        else:
            delta_value = "-"
            delta_help = "비교할 연도 데이터가 부족합니다"

    c1, c2, c3, c4, c5 = st.columns(5)
    total = fdf["금액"].sum()
    avg = period_totals.mean() if not period_totals.empty else 0
    peak = period_totals.idxmax() if not period_totals.empty else "-"
    n_firms = fdf["로펌"].nunique()

    c1.metric("총 비용 (VAT 제외)", f"₩{total:,.0f}",
              help=f"조회 기간({period_from} ~ {period_to}) 전체 합계")
    c2.metric(avg_label, f"₩{avg:,.0f}",
              help=f"{period_unit} 기준 평균 (구간 {n_periods}개)")
    c3.metric(peak_label, str(peak),
              help=f"{period_unit} 중 지출이 가장 컸던 구간")
    c4.metric("활성 로펌", f"{n_firms}곳",
              help="선택된 조건에서 지출이 발생한 로펌 수")
    c5.metric(delta_label, delta_value or "-", delta=delta_pct, help=delta_help)

    st.divider()

    # ---- 이번달 현황 (단일 연도 + 월별 집계에서만) — 2개월 페이지네이션 ----
    if view_mode == "단일 연도" and period_unit == "월별":
        # 전체 데이터 기준으로 최신 월 (필터와 무관, 인보이스가 있는 최신 시점)
        latest_period_global = sorted(df["표시기간"].unique(), reverse=True)[0] if not df.empty else None
        all_periods_desc = sorted(fdf["표시기간"].unique(), reverse=True)

        # 페이지네이션 offset (세션 상태)
        if "recent_offset" not in st.session_state:
            st.session_state["recent_offset"] = 0
        max_offset = max(0, len(all_periods_desc) - 2)
        offset = min(st.session_state["recent_offset"], max_offset)

        # 표시할 2개월
        display_months = all_periods_desc[offset:offset + 2]

        # 로펌 순서: 기본 알파벳 순 + 세종_인도네시아를 지평 뒤로
        firms_all = sorted(fdf["로펌"].unique())
        firms_sorted = [f for f in firms_all if f != "세종_인도네시아"]
        if "세종_인도네시아" in firms_all:
            firms_sorted.append("세종_인도네시아")

        # 제목 + 이전/다음 화살표
        col_prev, col_title, col_next = st.columns([1, 8, 1])
        with col_prev:
            if st.button("◀", disabled=offset >= max_offset, key="btn_prev_recent",
                         use_container_width=True):
                st.session_state["recent_offset"] = offset + 1
                st.rerun()
        with col_title:
            range_label = " · ".join(display_months) if display_months else "-"
            st.markdown(
                f"### 📋 이번달 현황  <span style='font-size:14px; color:#6c757d; font-weight:400;'>({range_label})</span>",
                unsafe_allow_html=True,
            )
        with col_next:
            if st.button("▶", disabled=offset <= 0, key="btn_next_recent",
                         use_container_width=True):
                st.session_state["recent_offset"] = offset - 1
                st.rerun()

        # HTML 표
        html = '<table style="width:100%; border-collapse:collapse; font-size:15px;">'
        html += '<tr style="background:#f1f3f5; border-bottom:2px solid #dee2e6;">'
        html += '<th style="padding:10px; text-align:center;">비용발생월</th>'
        for firm in firms_sorted:
            html += f'<th style="padding:10px; text-align:center;">{firm}</th>'
        html += '<th style="padding:10px; text-align:center; background:#dee2e6;">총 금액</th>'
        html += '</tr>'

        for period in display_months:
            is_latest = (period == latest_period_global)
            row_bg = "#FFF9C4" if is_latest else ""  # 연노랑
            row_style = f'background:{row_bg};' if row_bg else ''
            fw = "700" if is_latest else "500"
            html += f'<tr style="border-bottom:1px solid #dee2e6; {row_style}">'
            html += f'<td style="padding:10px; font-weight:{fw}; text-align:center;">{period}</td>'
            row_total = 0
            for firm in firms_sorted:
                row_data = fdf[(fdf["로펌"] == firm) & (fdf["표시기간"] == period)]
                if not row_data.empty:
                    amt = row_data["금액"].sum()
                    link = row_data.iloc[0]["링크"]
                    row_total += amt
                    if link:
                        html += f'<td style="padding:10px; text-align:center; font-weight:{fw};">'
                        html += f'<a href="{link}" target="_blank" style="color:#1B4F72; text-decoration:none;">₩{amt:,.0f}</a>'
                        html += '</td>'
                    else:
                        html += f'<td style="padding:10px; text-align:center; font-weight:{fw};">₩{amt:,.0f}</td>'
                else:
                    html += '<td style="padding:10px; text-align:center; color:#ccc;">-</td>'
            # 총 금액 컬럼
            html += (f'<td style="padding:10px; text-align:center; font-weight:700; '
                     f'color:#1B4F72; background:#E9ECEF;">₩{row_total:,.0f}</td>')
            html += '</tr>'
        html += '</table>'

        st.markdown(html, unsafe_allow_html=True)
        st.caption("💡 연노랑 = 전체 데이터 중 최신 월 · 금액 클릭 시 PDF 새 탭 · ◀▶로 이전/다음 월")
        st.markdown("<br>", unsafe_allow_html=True)
        st.divider()
    else:
        # 로펌 순서 (다른 섹션에서도 사용): 세종_인도네시아 뒤로
        firms_all = sorted(fdf["로펌"].unique())
        firms_sorted = [f for f in firms_all if f != "세종_인도네시아"]
        if "세종_인도네시아" in firms_all:
            firms_sorted.append("세종_인도네시아")

    # ---- 차트 ----
    sort_map = agg.groupby("집계")["__sort"].min().to_dict()
    periods_sorted = sorted(chart_df["집계"].unique(), key=lambda x: sort_map.get(x, 999))

    x_axis_title = {"월별": "월", "분기별": "분기", "반기별": "반기", "연도별": "연도"}[period_unit]

    if view_mode == "단일 연도" and chart_type == "비중 (도넛)":
        # ============ 도넛 차트 — 기준 기간 선택 ============
        # 집계 단위별 사용 가능한 기간 목록 (agg의 '집계' 컬럼)
        available_periods = sorted(chart_df["집계"].unique(), key=lambda x: sort_map.get(x, 999))
        # 셀렉트: 전체 기간 합계 + 각 기간
        donut_options = ["전체 기간 합계"] + available_periods
        # 기본값: 가장 최신 기간
        default_idx = len(donut_options) - 1
        donut_period = st.selectbox(
            "🍩 도넛 기준 기간",
            donut_options,
            index=default_idx,
            help="선택한 기간의 로펌별 비중을 표시. '전체 기간 합계'는 조회 기간 전체.",
        )

        # 데이터 필터
        if donut_period == "전체 기간 합계":
            donut_src = fdf
            title_period = f"{period_from} ~ {period_to}"
        else:
            donut_src = fdf.assign(__agg=fdf.apply(
                lambda r: build_agg_key(r, period_unit, cross_year=False), axis=1
            ))
            donut_src = donut_src[donut_src["__agg"] == donut_period]
            title_period = donut_period

        donut_df = (donut_src.groupby("로펌", as_index=False)["금액"].sum()
                             .query("금액 > 0")
                             .sort_values("금액", ascending=False))

        if donut_df.empty:
            st.warning(f"'{title_period}' 기간에 청구 내역이 없습니다.")
            fig = go.Figure()
        else:
            total_amt = donut_df["금액"].sum()
            fig = go.Figure()
            fig.add_trace(go.Pie(
                labels=donut_df["로펌"].tolist(),
                values=donut_df["금액"].tolist(),
                hole=0.45,
                marker=dict(colors=[FIRM_COLORS.get(f, "#95A5A6") for f in donut_df["로펌"]]),
                textinfo="label+percent",
                texttemplate="<b>%{label}</b><br>₩%{value:,.0f}<br>%{percent}",
                textfont=dict(size=13),
                hovertemplate="%{label}<br>₩%{value:,.0f} (%{percent})<extra></extra>",
                sort=False,
            ))
            fig.update_layout(
                title=dict(
                    text=f"<b>{title_period}</b> 로펌별 비중",
                    x=0.5, xanchor="center",
                    font=dict(size=15, color="#1B4F72"),
                ),
                annotations=[dict(
                    text=f"<b>총 ₩{total_amt/10_000:,.0f}만</b>",
                    x=0.5, y=0.5, font=dict(size=18, color="#1B4F72"), showarrow=False
                )],
                legend=dict(orientation="h", y=-0.05, x=0.5, xanchor="center", font=dict(size=13)),
                height=520,
                margin=dict(t=60, b=30, l=30, r=30),
                template="plotly_white",
            )
    elif view_mode == "단일 연도":
        # ============ 막대 차트 ============
        fig = go.Figure()
        for firm in sorted(chart_df["로펌"].unique()):
            fd = chart_df[chart_df["로펌"] == firm]
            fig.add_trace(
                go.Bar(
                    x=fd["집계"],
                    y=fd["금액"],
                    name=firm,
                    marker_color=FIRM_COLORS.get(firm, "#95A5A6"),
                    text=fd["금액"].apply(lambda v: f"{v/10_000:,.0f}만"),
                    textposition="inside",
                    textangle=0,
                    textfont=dict(size=14, color="white"),
                    hovertemplate="%{x}<br>%{fullData.name}: ₩%{y:,.0f}<extra></extra>",
                )
            )

        # ★ 카테고리별 총합 annotation — 최고 개별 막대 바로 위에 배치
        category_totals = chart_df.groupby("집계")["금액"].sum()
        category_max_bar = chart_df.groupby("집계")["금액"].max()  # 카테고리 안 최대 개별값
        annotations = []
        for cat in periods_sorted:
            tot = category_totals.get(cat, 0)
            max_bar = category_max_bar.get(cat, 0)
            if tot > 0:
                annotations.append(dict(
                    x=cat, y=max_bar,   # ← 총합 값이 아닌 '최고 개별 막대' 위치
                    text=f"<b>합계 ₩{tot/10_000:,.0f}만</b>",
                    showarrow=False,
                    yshift=20,
                    font=dict(size=15, color="#1B4F72"),
                ))

        # Y축 상한: 최대 개별 막대 기준 (총합 아님) — 데드스페이스 제거
        y_max = float(chart_df["금액"].max()) if not chart_df.empty else 0

        fig.update_layout(
            barmode="group",
            xaxis=dict(
                title=x_axis_title,
                type="category",
                categoryorder="array",
                categoryarray=periods_sorted,
                tickfont=dict(size=13),
            ),
            yaxis=dict(
                title="금액 (원)",
                tickformat=",",
                tickfont=dict(size=13),
                range=[0, y_max * 1.18] if y_max > 0 else None,
            ),
            legend=dict(orientation="h", y=-0.18, x=0.5, xanchor="center", font=dict(size=14)),
            height=520,
            margin=dict(t=30, b=80),
            template="plotly_white",
            hoverlabel=dict(font_size=14),
            annotations=annotations,
        )

        shapes = []
        for i in range(len(periods_sorted) - 1):
            shapes.append(dict(
                type="line", xref="x", yref="paper",
                x0=i + 0.5, x1=i + 0.5, y0=0, y1=1,
                line=dict(color="#ccc", width=1, dash="dot"),
            ))
        fig.update_layout(shapes=shapes)

    else:
        # ============ 라인 차트 (두 연도 비교) ============
        year_agg = fdf.groupby(["표시연도", "표시월"], as_index=False)["금액"].sum()
        year_agg["집계"] = year_agg.apply(
            lambda r: build_agg_key(r, period_unit, cross_year=True), axis=1
        )
        year_agg["__sort"] = year_agg.apply(sort_key, axis=1)
        year_chart = year_agg.groupby(["표시연도", "집계"], as_index=False).agg(
            금액=("금액", "sum"),
            __sort=("__sort", "min"),
        )
        periods_sorted = (
            year_chart[["집계", "__sort"]]
            .drop_duplicates()
            .sort_values("__sort")["집계"]
            .tolist()
        )

        year_palette = {2024: "#7BA7D9", 2025: "#E8998D", 2026: "#A3C9A8", 2027: "#B8A0D9"}

        fig = go.Figure()
        # 라벨 겹침 방지: 연도별로 textposition 다르게 (가장 작은 연도 = top, 그 외 = bottom)
        years_sorted = sorted(year_chart["표시연도"].unique())
        for idx, year in enumerate(years_sorted):
            yd = year_chart[year_chart["표시연도"] == year].sort_values("__sort")
            yd = yd.set_index("집계").reindex(periods_sorted).reset_index()
            # 짝수 번째 = top, 홀수 번째 = bottom (교대)
            tpos = "top center" if idx % 2 == 0 else "bottom center"
            fig.add_trace(
                go.Scatter(
                    x=yd["집계"],
                    y=yd["금액"],
                    mode="lines+markers+text",
                    name=f"{int(year)}년",
                    line=dict(color=year_palette.get(int(year), "#95A5A6"), width=3),
                    marker=dict(size=10),
                    text=yd["금액"].apply(
                        lambda v: f"{v/10_000:,.0f}만" if pd.notna(v) and v > 0 else ""
                    ),
                    textposition=tpos,
                    textfont=dict(size=14, color="#333"),
                    hovertemplate="%{x}<br>%{fullData.name}: ₩%{y:,.0f}<extra></extra>",
                    connectgaps=False,
                )
            )

        fig.update_layout(
            xaxis=dict(
                title=x_axis_title,
                type="category",
                categoryorder="array",
                categoryarray=periods_sorted,
                tickfont=dict(size=13),
            ),
            yaxis=dict(title="금액 (원)", tickformat=",", tickfont=dict(size=13)),
            legend=dict(orientation="h", y=-0.18, x=0.5, xanchor="center", font=dict(size=14)),
            height=520,
            margin=dict(t=30, b=80),
            template="plotly_white",
            hoverlabel=dict(font_size=14),
        )

    st.plotly_chart(fig, use_container_width=True)

    # ---- 피벗 테이블 (집계 단위 반영 + 합계 강조) ----
    st.subheader("📊 로펌별 기간 합계")

    # 집계 단위에 따른 컬럼 키 생성 (agg의 build_agg_key와 동일 로직)
    fdf_p = fdf.copy()
    fdf_p["_col"] = fdf_p.apply(
        lambda r: build_agg_key(r, period_unit, cross_year=False), axis=1
    )
    fdf_p["_sort"] = fdf_p.apply(sort_key, axis=1)

    pivot = fdf_p.pivot_table(
        values="금액",
        index="로펌",
        columns="_col",
        aggfunc="sum",
        fill_value=0,
        margins=True,
        margins_name="합계",
    )
    # 정렬 순서: sort_key 기준 (연도-월 순)
    col_order_map = fdf_p.groupby("_col")["_sort"].min().to_dict()
    # 두 연도 비교 모드에서 연도별 그룹핑 힌트도 반영
    if view_mode == "두 연도 비교" and period_unit != "연도별":
        # 연도-집계키 조합으로 정렬 (예: 2025-1월, 2026-1월, 2025-2월, ...)
        # 여기선 단순히 sort_key 순서 사용
        pass
    cols = sorted(
        [c for c in pivot.columns if c != "합계"],
        key=lambda c: col_order_map.get(c, 999),
    ) + ["합계"]
    pivot = pivot.reindex(columns=cols)

    # 두 연도 비교 모드일 때 월/분기/반기 그룹별로 배경색 교차
    def _group_key(c):
        if not isinstance(c, str) or c == "합계":
            return c
        parts = c.split("-", 1)
        return parts[1] if len(parts) == 2 else c

    data_cols_ordered = [c for c in pivot.columns if c != "합계"]
    col_group_bg = {}
    if view_mode == "두 연도 비교" and len(data_cols_ordered) > 1:
        groups_seen = []
        for c in data_cols_ordered:
            g = _group_key(c)
            if g not in groups_seen:
                groups_seen.append(g)
        for c in data_cols_ordered:
            g = _group_key(c)
            idx = groups_seen.index(g)
            col_group_bg[c] = "background-color: #F5F7FA;" if idx % 2 == 1 else ""

    def _highlight_and_group(df):
        styles = pd.DataFrame("", index=df.index, columns=df.columns)
        emphasis = "background-color: #E9ECEF; font-weight: 700; color: #1B4F72;"
        for c, bg in col_group_bg.items():
            if bg and c in df.columns:
                for r in df.index:
                    if r != "합계":
                        styles.loc[r, c] = bg
        if "합계" in df.index:
            for c in df.columns:
                styles.loc["합계", c] = emphasis
        if "합계" in df.columns:
            for r in df.index:
                styles.loc[r, "합계"] = emphasis
        return styles

    st.dataframe(
        pivot.style
            .format("₩{:,.0f}")
            .map(lambda v: "color: #ccc" if v == 0 else "", subset=pivot.columns)
            .apply(_highlight_and_group, axis=None),
        use_container_width=True,
    )

    # ---- 상세 내역 ----
    st.subheader("📋 상세 내역")

    fc1, fc2 = st.columns(2)
    with fc1:
        detail_firms = st.multiselect(
            "로펌 필터", firms_sorted, default=firms_sorted, key="detail_firm"
        )
    with fc2:
        all_months = sorted(fdf["표시기간"].unique(), reverse=True)
        detail_months = st.multiselect(
            "월 필터", all_months, default=all_months, key="detail_month"
        )

    tbl = fdf[
        (fdf["로펌"].isin(detail_firms)) & (fdf["표시기간"].isin(detail_months))
    ][["로펌", "표시기간", "금액", "파일명", "링크"]].copy()
    tbl = tbl.sort_values(["표시기간", "로펌"], ascending=[False, True]).reset_index(drop=True)
    tbl["금액(VAT제외)"] = tbl["금액"].apply(lambda x: f"₩{x:,.0f}")
    tbl = tbl.drop(columns=["금액"])
    tbl = tbl.rename(columns={"표시기간": "비용발생월"})

    st.dataframe(
        tbl,
        column_config={
            "링크": st.column_config.LinkColumn("PDF", display_text="📄 열기"),
            "파일명": st.column_config.TextColumn("파일명", width="large"),
        },
        hide_index=True,
        use_container_width=True,
    )


# ============================================================
# 실행
# ============================================================
if __name__ == "__main__":
    main()
